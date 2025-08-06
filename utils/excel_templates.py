# -*- coding: utf-8 -*-
"""
Gerador de Templates Excel para Marmitas Fit
Templates para ingredientes, embalagens e custos fixos
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import streamlit as st

class ExcelTemplateGenerator:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def create_ingredientes_template(self):
        """Cria template para ingredientes"""
        
        # Dados de exemplo
        data = [
            {
                'Nome': 'Frango (peito)',
                'Categoria': 'Proteína Animal',
                'Preço (R$)': 18.90,
                'Unid.Receita': 'g',
                'Unid.Compra': 'kg',
                'Kcal/Unid': 1.65,
                'Fator Conv.': 1000,
                'Ativo': True,
                'Observações': 'Sem pele, congelado'
            },
            {
                'Nome': 'Arroz integral',
                'Categoria': 'Carboidrato',
                'Preço (R$)': 8.90,
                'Unid.Receita': 'g',
                'Unid.Compra': 'kg',
                'Kcal/Unid': 1.11,
                'Fator Conv.': 1000,
                'Ativo': True,
                'Observações': 'Grão longo, tipo 1'
            },
            {
                'Nome': 'Brócolis',
                'Categoria': 'Vegetal',
                'Preço (R$)': 6.50,
                'Unid.Receita': 'g',
                'Unid.Compra': 'kg',
                'Kcal/Unid': 0.34,
                'Fator Conv.': 1000,
                'Ativo': True,
                'Observações': 'Fresco, preço médio'
            },
            {
                'Nome': 'Azeite extra virgem',
                'Categoria': 'Gordura',
                'Preço (R$)': 12.00,
                'Unid.Receita': 'ml',
                'Unid.Compra': 'L',
                'Kcal/Unid': 8.84,
                'Fator Conv.': 1000,
                'Ativo': True,
                'Observações': 'Primeira prensagem'
            },
            {
                'Nome': 'Sal refinado',
                'Categoria': 'Tempero',
                'Preço (R$)': 1.20,
                'Unid.Receita': 'g',
                'Unid.Compra': 'kg',
                'Kcal/Unid': 0.00,
                'Fator Conv.': 1000,
                'Ativo': True,
                'Observações': 'Iodado'
            }
        ]
        
        return self._create_excel_from_data(
            data, 
            "Template Ingredientes - Marmitas Fit",
            "Preencha com os ingredientes disponíveis no mercado"
        )
    
    def create_embalagens_template(self):
        """Cria template para embalagens"""
        
        data = [
            {
                'Nome': 'Marmita 500ml',
                'Tipo': 'descartavel',
                'Preço (R$)': 0.50,
                'Capacidade (ml)': 500,
                'Categoria': 'principal',
                'Ativo': True,
                'Descrição': 'PP transparente com tampa'
            },
            {
                'Nome': 'Marmita 750ml',
                'Tipo': 'descartavel',
                'Preço (R$)': 0.65,
                'Capacidade (ml)': 750,
                'Categoria': 'principal',
                'Ativo': True,
                'Descrição': 'PP transparente com tampa'
            },
            {
                'Nome': 'Marmita 1000ml',
                'Tipo': 'descartavel',
                'Preço (R$)': 0.80,
                'Capacidade (ml)': 1000,
                'Categoria': 'principal',
                'Ativo': True,
                'Descrição': 'PP transparente com tampa'
            },
            {
                'Nome': 'Pote sobremesa 150ml',
                'Tipo': 'descartavel',
                'Preço (R$)': 0.25,
                'Capacidade (ml)': 150,
                'Categoria': 'complemento',
                'Ativo': True,
                'Descrição': 'Para doces e frutas'
            },
            {
                'Nome': 'Talher plástico',
                'Tipo': 'utensilio',
                'Preço (R$)': 0.08,
                'Capacidade (ml)': 0,
                'Categoria': 'utensilio',
                'Ativo': True,
                'Descrição': 'Garfo + faca + colher'
            },
            {
                'Nome': 'Guardanapo',
                'Tipo': 'higiene',
                'Preço (R$)': 0.05,
                'Capacidade (ml)': 0,
                'Categoria': 'higiene',
                'Ativo': True,
                'Descrição': 'Papel 20x20cm'
            },
            {
                'Nome': 'Sacola plástica',
                'Tipo': 'transporte',
                'Preço (R$)': 0.12,
                'Capacidade (ml)': 0,
                'Categoria': 'transporte',
                'Ativo': True,
                'Descrição': '30x40cm alça camiseta'
            }
        ]
        
        return self._create_excel_from_data(
            data,
            "Template Embalagens - Marmitas Fit",
            "Defina os custos das embalagens utilizadas"
        )
    
    def create_custos_fixos_template(self):
        """Cria template para custos fixos"""
        
        data = [
            {
                'Categoria': 'Energia',
                'Item': 'Conta de luz',
                'Custo Mensal (R$)': 150.00,
                'Rateio por Marmita': 0.30,
                'Descrição': 'Fogão, geladeira, freezer'
            },
            {
                'Categoria': 'Gás',
                'Item': 'Botijão 13kg',
                'Custo Mensal (R$)': 80.00,
                'Rateio por Marmita': 0.16,
                'Descrição': 'Consumo médio mensal'
            },
            {
                'Categoria': 'Água',
                'Item': 'Conta de água',
                'Custo Mensal (R$)': 60.00,
                'Rateio por Marmita': 0.12,
                'Descrição': 'Limpeza e preparo'
            },
            {
                'Categoria': 'Aluguel',
                'Item': 'Espaço cozinha',
                'Custo Mensal (R$)': 800.00,
                'Rateio por Marmita': 1.60,
                'Descrição': 'Proporcional ao uso'
            },
            {
                'Categoria': 'Mão de obra',
                'Item': 'Salário próprio',
                'Custo Mensal (R$)': 2000.00,
                'Rateio por Marmita': 4.00,
                'Descrição': 'Base: 500 marmitas/mês'
            },
            {
                'Categoria': 'TOTAL',
                'Item': '',
                'Custo Mensal (R$)': 3090.00,
                'Rateio por Marmita': 6.18,
                'Descrição': 'Base: 500 marmitas/mês'
            }
        ]
        
        return self._create_excel_from_data(
            data,
            "Template Custos Fixos - Marmitas Fit",
            "Calcule seus custos fixos mensais por marmita"
        )
    
    def _create_excel_from_data(self, data, title, subtitle):
        """Cria arquivo Excel formatado a partir dos dados"""
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Título
        ws.merge_cells('A1:' + chr(65 + len(data[0]) - 1) + '1')
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color="2E7D32")
        ws['A1'].alignment = self.center_alignment
        
        # Subtítulo
        ws.merge_cells('A2:' + chr(65 + len(data[0]) - 1) + '2')
        ws['A2'] = subtitle
        ws['A2'].font = Font(size=12, color="666666")
        ws['A2'].alignment = self.center_alignment
        
        # Espaço
        ws.append([])
        
        # Converter para DataFrame
        df = pd.DataFrame(data)
        
        # Adicionar headers
        headers = list(df.columns)
        ws.append(headers)
        
        # Formatar headers
        header_row = ws.max_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Adicionar dados
        for row_data in dataframe_to_rows(df, index=False, header=False):
            ws.append(row_data)
        
        # Formatar dados
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            for cell in row:
                cell.border = self.border
                if isinstance(cell.value, bool):
                    cell.alignment = self.center_alignment
                elif isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()

# Funções para uso no Streamlit
def generate_ingredientes_template():
    """Gera template de ingredientes para download"""
    generator = ExcelTemplateGenerator()
    return generator.create_ingredientes_template()

def generate_embalagens_template():
    """Gera template de embalagens para download"""
    generator = ExcelTemplateGenerator()
    return generator.create_embalagens_template()

def generate_custos_fixos_template():
    """Gera template de custos fixos para download"""
    generator = ExcelTemplateGenerator()
    return generator.create_custos_fixos_template()